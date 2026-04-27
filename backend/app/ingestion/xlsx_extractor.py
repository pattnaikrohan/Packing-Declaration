"""
XLSX extractor using openpyxl.
Scans all sheets for yes/no/✓/✗ markers near question labels.
"""
from openpyxl import load_workbook
from io import BytesIO

from app.ingestion.schema import PackingDeclaration
from app.ingestion import extractors_common as ec

YES_MARKERS = {"yes", "y", "✓", "x", "☑", "☒", "true", "1"}
NO_MARKERS = {"no", "n", "✗", "☐", "false", "0"}


def _cell_val(cell) -> str:
    if cell.value is None:
        return ""
    return str(cell.value).strip().lower()


def _is_yes(val: str) -> bool:
    return val in YES_MARKERS


def _is_no(val: str) -> bool:
    return val in NO_MARKERS


def extract(file_bytes: bytes) -> PackingDeclaration:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    all_text_parts = []
    checkboxes = {"q1": "BLANK", "q2": "BLANK", "q3": "BLANK", "q4": "ABSENT"}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows())
        for row in rows:
            for col_idx, cell in enumerate(row):
                val = _cell_val(cell)
                if not val:
                    continue

                all_text_parts.append(str(cell.value or ""))
                upper = val.upper()

                answer_cell = ""
                if col_idx + 1 < len(row):
                    answer_cell = _cell_val(row[col_idx + 1])
                if not answer_cell and col_idx + 2 < len(row):
                    answer_cell = _cell_val(row[col_idx + 2])

                if "Q1" in upper or "UNACCEPTABLE" in upper or "PROHIBIT" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q1"] = "YES"
                    elif _is_no(answer_cell) or _is_no(val):
                        checkboxes["q1"] = "NO"

                elif "TIMBER" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q2"] = "YES_TIMBER"
                    elif _is_no(answer_cell) or _is_no(val):
                        if checkboxes["q2"] == "BLANK":
                            checkboxes["q2"] = "NO"

                elif "BAMBOO" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q2"] = "YES_BAMBOO"

                elif "Q2" in upper:
                    if _is_yes(answer_cell):
                        if checkboxes["q2"] == "BLANK":
                            checkboxes["q2"] = "YES_TIMBER"
                    elif _is_no(answer_cell):
                        if checkboxes["q2"] == "BLANK":
                            checkboxes["q2"] = "NO"

                elif "ISPM" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q3"] = "ISPM15"

                elif "DAFF" in upper and "CERTIF" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q3"] = "DAFF_CERTIFIED"

                elif "NOT TREATED" in upper or "UNTREATED" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q3"] = "NOT_TREATED"

                elif ("NOT APPLICABLE" in upper or "N/A" in upper):
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q3"] = "NOT_APPLICABLE"

                elif "Q3" in upper or "TREATMENT" in upper:
                    if _is_yes(answer_cell) and checkboxes["q3"] == "BLANK":
                        checkboxes["q3"] = "ISPM15"
                    elif _is_no(answer_cell) and checkboxes["q3"] == "BLANK":
                        checkboxes["q3"] = "NOT_APPLICABLE"

                elif "Q4" in upper or "CLEANLINESS" in upper or "CLEAN" in upper:
                    if _is_yes(answer_cell) or _is_yes(val):
                        checkboxes["q4"] = "PRESENT"
                    elif _is_no(answer_cell) or _is_no(val):
                        checkboxes["q4"] = "ABSENT"

    full_text = "\n".join(all_text_parts)

    addr, is_po = ec.extract_address(full_text)
    consignment, _ = ec.extract_consignment_link(full_text)
    date_str, date_valid = ec.extract_date(full_text)
    signed, sig_type = ec.detect_signature(full_text)
    alterations, endorsed = ec.detect_alterations(full_text)

    return PackingDeclaration(
        declaration_type=ec.detect_declaration_type(full_text),
        issuer_company=ec.extract_company(full_text),
        issuer_address=addr,
        issuer_address_is_po_box=is_po,
        vessel_name=ec.extract_vessel(full_text),
        voyage_number=ec.extract_voyage(full_text),
        consignment_ref=consignment,
        exporter=ec.extract_party(full_text, "Exporter"),
        importer=ec.extract_party(full_text, "Importer"),
        date_issued=date_str,
        date_valid=date_valid,
        signed=signed,
        signature_type=sig_type,
        printed_name=ec.extract_printed_name(full_text),
        letterhead_present=ec.detect_letterhead(full_text),
        q1_unacceptable_material=checkboxes["q1"],
        q2_timber_bamboo=checkboxes["q2"],
        q3_treatment=checkboxes["q3"],
        q4_cleanliness=checkboxes["q4"],
        alterations_present=alterations,
        alterations_endorsed=endorsed,
        extraction_method="xlsx",
        ocr_confidence=1.0,
    )
